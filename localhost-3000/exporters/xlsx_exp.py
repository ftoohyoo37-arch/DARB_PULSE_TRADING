from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from auditor.models import Finding


SEV_FILLS = {
    "error":   PatternFill("solid", fgColor="F8D7DA"),
    "warning": PatternFill("solid", fgColor="FFF3CD"),
    "info":    PatternFill("solid", fgColor="D1ECF1"),
}


def to_xlsx(meta: dict[str, Any], findings: list[Finding]) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Crossref Audit Report"
    summary["A1"].font = Font(bold=True, size=16)
    rows = [
        ("File", meta.get("filename", "")),
        ("File size (bytes)", meta.get("file_size", 0)),
        ("Run at", meta.get("created_at", "")),
        ("Schema namespace", meta.get("namespace") or ""),
        ("Citations checked", meta.get("citation_n", 0)),
        ("Errors", meta.get("error_n", 0)),
        ("Warnings", meta.get("warning_n", 0)),
        ("Info", meta.get("info_n", 0)),
        ("Total findings", len(findings)),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        summary.cell(row=i, column=1, value=k).font = Font(bold=True)
        summary.cell(row=i, column=2, value=v)
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 60

    findings_ws = wb.create_sheet("Findings")
    headers = ["Severity", "Rule", "Line", "Citation key", "Message", "Snippet", "XPath"]
    for c, h in enumerate(headers, start=1):
        cell = findings_ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")
    for r, f in enumerate(findings, start=2):
        values = [f.severity, f.rule_id, f.line, f.citation_key, f.message, f.snippet, f.xpath]
        for c, v in enumerate(values, start=1):
            cell = findings_ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        fill = SEV_FILLS.get(f.severity)
        if fill:
            findings_ws.cell(row=r, column=1).fill = fill
    widths = [10, 24, 8, 16, 60, 60, 30]
    for i, w in enumerate(widths, start=1):
        findings_ws.column_dimensions[get_column_letter(i)].width = w
    findings_ws.freeze_panes = "A2"
    findings_ws.auto_filter.ref = findings_ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
